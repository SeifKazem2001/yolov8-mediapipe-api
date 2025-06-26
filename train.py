# Import required libraries
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

# Create workbook
wb = Workbook()
ws = wb.active

# Set column headers with Date column added
headers = ['Date', 'Start_Time', 'End_Time', 'Exercise_Type', 'Sets', 'Reps_Per_Set', 'Total_Reps', 'Notes']
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# Auto-populate the Date column with today's date and month name using Excel formula
# This will display date in format like "May 01" if today's date is May 1
# Use formula TEXT(TODAY(),"mmmm dd") in cell A2 and fill down
ws['A2'] = '=TEXT(TODAY(),"mmmm dd")'
for row in range(3, 1001):
    ws.cell(row=row, column=1, value='=TEXT(TODAY(),"mmmm dd")')

# Create list of exercises for dropdown
exercises = [
    'Pull-ups',
    'Push-ups',
    'Bicep Body Weight Curls',
    'Squats',
    'Lunges',
    'Dips',
    'Plank',
    'Mountain Climbers',
    'Burpees',
    'Chin-ups'
]

# Create data validation for Exercise_Type column (now column D)
dv = DataValidation(type="list", formula1='"' + ','.join(exercises) + '"', allow_blank=True)
ws.add_data_validation(dv)
dv.add('D2:D1000')

# Define fills for conditional formatting
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

# Green rule: if cell in column D (Exercise_Type) is not blank
green_rule = FormulaRule(formula=['NOT(ISBLANK($D2))'], fill=green_fill)

# Red rule: if cell in column D is blank
red_rule = FormulaRule(formula=['ISBLANK($D2)'], fill=red_fill)

# Apply the rules for the entire range A2:H1000
ws.conditional_formatting.add('A2:H1000', green_rule)
ws.conditional_formatting.add('A2:H1000', red_rule)

# Save workbook to file
wb.save('workout_tracker_date.xlsx')

# Create example DataFrame to show structure for reference
sample_data = {
    'Date': ['May 01', 'May 01', 'May 01'],
    'Start_Time': ['07:00', '07:15', '07:30'],
    'End_Time': ['07:15', '07:30', '07:45'],
    'Exercise_Type': ['Pull-ups', 'Push-ups', ''],
    'Sets': [3, 4, ''],
    'Reps_Per_Set': ['8,8,7', '15,12,12,10', ''],
    'Total_Reps': [23, 49, ''],
    'Notes': ['Good form', 'Felt strong', 'No workout']
}
example_df = pd.DataFrame(sample_data)
print("Workout tracker structure with Date column:")
print(example_df)

print("Excel file 'workout_tracker_date.xlsx' created with auto date and month formatting in column A.")